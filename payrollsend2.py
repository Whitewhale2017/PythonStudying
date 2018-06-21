# -*- coding: utf-8 -*-
"""
@title = '工资邮件发送'
"""
import datetime
import shutil
import os
import logging
import logging.config

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.utils import parseaddr, formataddr

from openpyxl import load_workbook

EMAIL_HOST = 'mail.chyjr.com'
EMAIL_PORT = 465
EMAIL_HOST_USER = 'payrollsend@chyjr.com'
EMAIL_HOST_PASSWORD = '1q2w3e4R'


def get_log_config():
    _config = {
        'version': 1,
        'formatters': {
            'generic': {
                'format': '%(asctime)s %(levelname)-5.5s [%(name)s:%(lineno)s][%(threadName)s] %(message)s',
            },
            'simple': {
                'format': '%(asctime)s %(levelname)-5.5s %(message)s',
            },
        },
        'handlers': {
            'console': {
                'class': 'logging.StreamHandler',
                'formatter': 'generic',
            },
            'file': {
                'class': 'logging.FileHandler',
                'filename': "payrollsend.log",
                'encoding': 'utf-8',
                'formatter': 'generic',

            },
        },
        'root': {
            'level': "INFO",
            'handlers': ['console', 'file', ],
        }
    }
    return _config


log_config = get_log_config()
logging.config.dictConfig(log_config)
logger = logging.getLogger(__file__)


_COL_I = {
    'A': 0,
    'B': 1,
    'C': 2,
    'D': 3,
    'E': 4,
    'F': 5,
    'G': 6,
    'H': 7,
    'I': 8,
    'J': 9,
    'K': 10,
    'L': 11,
    'M': 12,
    'N': 13,
    'O': 14,
    'P': 15,
    'Q': 16,
    'R': 17,
    'S': 18,
    'T': 19,
    'U': 20,
    'V': 21,
    'W': 22,
    'X': 23,
    'Y': 24,
    'Z': 25,
    'AA': 26,
    'AB': 27,
    'AC': 28,
    'AD': 29,
    'AE': 30,
    'AF': 31,
    'AG': 32,
    'AH': 33,
    'AI': 34,
    'AJ': 35,
    'AK': 36,
    'AL': 37,
    'AM': 38,
    'AN': 39,
    'AO': 40,
    'AP': 41,
    'AQ': 42,
    'AR': 43,
    'AS': 44,
    'AT': 45,
    'AU': 46,
    'AV': 47,
    'AW': 48,
    'AX': 49,
    'AY': 50,
    'AZ': 51,
}

COL_MAX = _COL_I['AG'] + 1


def _format_float(obj):
    if isinstance(obj, float):
        return '{:.2f}'.format(obj)
    else:
        return str(obj)


def _format_percent(obj):
    if isinstance(obj, float) or isinstance(obj, int):
        t = obj*100
        return '{:.0f}'.format(t) + '%'
    else:
        return str(obj)


def _format_int(obj):
    if isinstance(obj, float):
        return '{:.0f}'.format(obj)
    else:
        return str(obj)


def _format(obj):
    if obj is None:
        return ''
    if isinstance(obj, datetime.datetime):
        # return obj.strftime('%Y-%m-%d %H:%M:%S')
        return obj.strftime('%Y-%m-%d')
    elif isinstance(obj, datetime.date):
        return obj.strftime('%Y-%m-%d')
    elif isinstance(obj, str):
        return str(obj).strip()
    else:
        # print(type(obj))
        # print(obj)
        return obj


def get_datetime_str(d_date=None, pattern='%Y-%m-%d'):
    """
    获取指定日期 字符格式
    :param d_date:
    :param pattern:
    :return:
    """
    if not d_date:
        d_date = datetime.datetime.now()
    return datetime.datetime.strftime(d_date, pattern)


def read_xlsx(file_name):
    """
    读xlsx
    :param file_name:
    :return:
    """
    logger.info('读取数据  %s' % file_name)
    res = []
    wb = load_workbook(file_name)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    key_index = 0
    for row in sheet.rows:
        key_index += 1
        if key_index >= 2:
            if row[0].value is None or row[0].value == '':
                continue
            if len(row) > COL_MAX:
                row = row[:COL_MAX]
            values = [_format(c.value) for c in row]
            res.append(values)
    return res


def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))


def _get_html_1(data):
    html_message = '''
            <table border="1"  style="border-collapse:collapse;">
                <tr>
                    <td>工号</td>
                    <td>机构</td>
                    <td>姓名</td>
                    <td>邮箱</td>
                    <td>基本工资</td>
                    <td>未入职或离职</td>
                    <td>扣病事假或过失</td>
                    <td>其他应扣</td>
                    <td>应发基本工资</td>
                    <td>代扣养老保险</td>
                    <td>代扣医疗保险</td>
                    <td>代扣失业保险</td>
                    <td>代扣公积金</td>
                    <td>代扣个税</td>
                    <td>实发基本工资</td>
                    <td>*</td>
                    <td>绩效工资</td>
                    <td>考勤扣款</td>
                    <td>应发绩效工资</td>
                    <td>绩效考核1：月度绩效奖金</td>
                    <td>绩效考核2：月度绩效工资</td>
                    <td>绩效考核3：专项奖罚</td>
                    <td>绩效考核4：招聘奖罚</td>
                    <td>绩效考核5：</td>
                    <td>绩效考核6：</td>
                    <td>交通通讯津贴</td>
                    <td>工龄工资</td>
                    <td>应发工作餐</td>
                    <td>其他应加应扣</td>
                    <td>补税</td>
                    <td>实发绩效工资</td>
                    <td>工资合计</td>
                    <td>备注</td>
                </tr>
                <tr>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                    <td>%s</td>
                </tr>
            </table>
            ''' % (str(data[_COL_I['A']]),
                   str(data[_COL_I['B']]),
                   str(data[_COL_I['C']]),
                   str(data[_COL_I['AG']]),
                   _format_float(data[_COL_I['D']]),
                   _format_float(data[_COL_I['E']]),
                   _format_float(data[_COL_I['F']]),
                   _format_float(data[_COL_I['G']]),
                   _format_float(data[_COL_I['H']]),
                   _format_float(data[_COL_I['I']]),
                   _format_float(data[_COL_I['J']]),
                   _format_float(data[_COL_I['K']]),
                   _format_float(data[_COL_I['L']]),
                   _format_float(data[_COL_I['M']]),
                   _format_float(data[_COL_I['N']]),
                   '*',
                   _format_float(data[_COL_I['P']]),
                   _format_float(data[_COL_I['Q']]),
                   _format_float(data[_COL_I['R']]),
                   _format_float(data[_COL_I['S']]),
                   _format_float(data[_COL_I['T']]),
                   _format_float(data[_COL_I['U']]),
                   _format_float(data[_COL_I['V']]),
                   _format_float(data[_COL_I['W']]),
                   _format_float(data[_COL_I['X']]),
                   _format_float(data[_COL_I['Y']]),
                   _format_float(data[_COL_I['Z']]),
                   _format_float(data[_COL_I['AA']]),
                   _format_float(data[_COL_I['AB']]),
                   _format_float(data[_COL_I['AC']]),
                   _format_float(data[_COL_I['AD']]),
                   _format_float(data[_COL_I['AE']]),
                   str(data[_COL_I['AF']]),

                   )
    return html_message, data[_COL_I['AG']]


def send_stub(datas):
    """
    发送时间
    :param datas:
    :return:
    """
    logging.info("邮件发送......")
    l_month = datetime.date.today().replace(day=1) - datetime.timedelta(1)
    subject = str(l_month.year) + '年' + str(l_month.month) + '月工资条'
    for data in datas:
        server = None
        try:
            server = smtplib.SMTP_SSL('mail.chyjr.com', 465)
            server.ehlo()
            server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
            try:
                html_message, receiver = _get_html_1(data)
                # print(html_message)
                # print(receiver)
                msg = MIMEMultipart()
                msg['Subject'] = subject
                msg['To'] = receiver
                msg_text = MIMEText(html_message, 'html', 'utf-8')
                msg.attach(msg_text)
                server.sendmail(EMAIL_HOST_USER, receiver, msg.as_string())
                data.append('发送成功')
                logger.info(data)
            except Exception as e:
                data.append('发送失败')
                data.append(str(e))
                logger.error(data)
        finally:
            try:
                if server:
                    server.quit()
            except Exception as e:
                logging.info(str(e))


def _write_cell(wb, res, start_row):
    """

    :param wb:
    :param res:
    :return:
    """
    if len(res) > 0:
        sheetnames = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheetnames[0])
        row_index = start_row
        for r in res:
            if len(r) > COL_MAX:
                for i in range(COL_MAX, len(r)):
                    sheet.cell(row=row_index, column=i+1, value=r[i])
            row_index += 1


def insert_send_result(file_name, res1):
    """
    记录发送结果
    :param file_name:
    :param res1:
    :return:
    """
    logging.info("记录发送结果...")
    wb = load_workbook(file_name)
    _write_cell(wb, res1, 2)
    wb.save(file_name)


def main():
    """
    入口
    :return:
    """
    func_name = "工资邮件发送"
    logger.info("start %s" % func_name)
    d_now = datetime.datetime.now()
    s_now = d_now.strftime("%Y%m%d%H%M%S")
    # 当前路径
    pwd_path = os.getcwd()
    # 文件存放路径
    cur_path = os.path.join(pwd_path, 'src')
    cur_dirs = os.listdir(cur_path)
    file_name = ''
    for cur_dir in cur_dirs:
        if cur_dir.find('.xlsx') != -1:
            file_name = cur_dir
    if not file_name:
        logging.info("没有Excel文件!")
        return
    logging.info(file_name)
    full_name = os.path.join(cur_path, file_name)
    # 将文件拷贝到desc目录
    desc_full = os.path.join(os.path.join(pwd_path, 'desc'), s_now + file_name)
    shutil.copy(full_name, desc_full)

    res1 = read_xlsx(full_name)
    send_stub(res1)

    insert_send_result(desc_full, res1)
    # 删除src目录内容
    os.remove(full_name)
    logger.info('... end %s' % func_name)


if __name__ == '__main__':
    main()





